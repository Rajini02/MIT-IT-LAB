import openpyxl
import math
# load excel with its path
yc=0
nc=0
info_d=0
max1=0

o_s_n=0
o_s_y=0
o_o_n=0
o_o_y=0
o_r_y=0
o_r_n=0
info_out=0
infogain_out=0

t_h_y=0
t_h_n=0
t_m_y=0
t_m_n=0
t_c_y=0
t_c_n=0
info_temp=0
infogain_temp=0

h_h_y=0
h_h_n=0
h_n_y=0
h_n_n=0
info_hum=0
infogain_hum=0

w_s_y=0
w_s_n=0
w_w_y=0
w_w_n=0
info_wind=0
infogain_wind=0

wrkbk = openpyxl.load_workbook("/Users/PRANAV/Desktop/VI sem/dwdm lab/id3/data.xlsx")
sh=wrkbk.active
#for row in range(0, sh.max_row):
#    for col in sh.iter_cols(1, sh.max_column):
        #print(col[row].value,end=' ')
    #print('')

for row in range(1,sh.max_row+1):
    #col=6
    cellobj=sh.cell(row=row,column=6)
    #print(cellobj.value)
    if cellobj.value == "Yes":
        yc += 1
    else:
        nc += 1
#print("Yes count = {}".format(yc))
#print("No count = {}".format(nc))

info_d=((-yc/(yc+nc))*math.log((yc)/(yc+nc),2))-((nc/(yc+nc))*math.log((nc)/(yc+nc),2))
#print(info_d)

for row in range(1,sh.max_row+1):
    #col=6
    cellobj=sh.cell(row=row,column=2)
    cellobj2=sh.cell(row=row,column=6)
    #print(cellobj.value)
    if cellobj.value == "Sunny" and cellobj2.value == "Yes":
        o_s_y += 1
    elif cellobj.value == "Sunny" and cellobj2.value == "No":
        o_s_n +=1
    elif cellobj.value == "Rain" and cellobj2.value == "Yes":
        o_r_y += 1
    elif cellobj.value == "Rain" and cellobj2.value == "No":
        o_r_n += 1
    elif cellobj.value == "Overcast" and cellobj2.value == "No":
        o_o_n += 1
    elif cellobj.value == "Overcast" and cellobj2.value == "Yes":
        o_o_y += 1
#print(o_s_y,o_s_n,o_r_y,o_r_n,o_o_y,o_o_n)
info_out = (((o_s_y+o_s_n)/(yc+nc))*(((-o_s_y/(o_s_y+o_s_n))*math.log(o_s_y/(o_s_y+o_s_n),2))+((-o_s_n/(o_s_y+o_s_n))*math.log(o_s_n/(o_s_y+o_s_n),2))))
+ (((o_r_y+o_r_n)/(yc+nc))*(((-o_r_y/(o_r_y+o_r_n))*math.log(o_r_y/(o_r_y+o_r_n),2))+((-o_r_n/(o_r_y+o_r_n))*math.log(o_r_n/(o_r_y+o_r_n),2))))
+ (((o_o_y+o_o_n)/(yc+nc))*(((-o_o_y/(o_o_y+o_o_n))*math.log(o_o_y/(o_o_y+o_o_n),2))))

#print(info_out)
infogain_out = info_d - info_out
#print(infogain_out)

for row in range(1,sh.max_row+1):
    #col=6
    cellobj=sh.cell(row=row,column=3)
    cellobj2=sh.cell(row=row,column=6)
    #print(cellobj.value)
    if cellobj.value == "Hot" and cellobj2.value == "Yes":
        t_h_y += 1
    elif cellobj.value == "Hot" and cellobj2.value == "No":
        t_h_n +=1
    elif cellobj.value == "Mild" and cellobj2.value == "Yes":
        t_m_y += 1
    elif cellobj.value == "Mild" and cellobj2.value == "No":
        t_m_n += 1
    elif cellobj.value == "Cool" and cellobj2.value == "No":
        t_c_n += 1
    elif cellobj.value == "Cool" and cellobj2.value == "Yes":
        t_c_y += 1
#print(t_h_y,t_h_n,t_m_y,t_m_n,t_c_y,t_c_n)

info_temp = (((t_h_y+t_h_n)/(yc+nc))*(((-t_h_y/(t_h_y+t_h_n))*math.log(t_h_y/(t_h_y+t_h_n),2))+((-t_h_n/(t_h_y+t_h_n))*math.log(t_h_n/(t_h_y+t_h_n),2))))
+ (((t_m_y+t_m_n)/(yc+nc))*(((-t_m_y/(t_m_y+t_m_n))*math.log(t_m_y/(t_m_y+t_m_n),2))+((-t_m_n/(t_m_y+t_m_n))*math.log(t_m_n/(t_m_y+t_m_n),2))))
+ (((t_c_y+t_c_n)/(yc+nc))*(((-t_c_y/(t_c_y+t_c_n))*math.log(t_c_y/(t_c_y+t_c_n),2))+((-t_c_n/(t_c_y+t_c_n))*math.log(t_c_n/(t_c_y+t_c_n),2))))
#print(info_temp)
infogain_temp = info_d-info_temp
#print(infogain_temp)

for row in range(1,sh.max_row+1):
    #col=6
    cellobj=sh.cell(row=row,column=4)
    cellobj2=sh.cell(row=row,column=6)
    #print(cellobj.value)
    if cellobj.value == "High" and cellobj2.value == "Yes":
        h_h_y += 1
    elif cellobj.value == "High" and cellobj2.value == "No":
        h_h_n +=1
    elif cellobj.value == "Normal" and cellobj2.value == "Yes":
        h_n_y += 1
    elif cellobj.value == "Normal" and cellobj2.value == "No":
        h_n_n += 1
#print(h_h_y,h_h_n,h_n_y,h_n_n)

info_hum = (((h_h_y+h_h_n)/(yc+nc))*(((-h_h_y/(h_h_y+h_h_n))*math.log(h_h_y/(h_h_y+h_h_n),2))+((-h_h_n/(h_h_y+h_h_n))*math.log(h_h_n/(h_h_y+h_h_n),2))))
+ (((h_n_y+h_n_n)/(yc+nc))*(((-h_n_y/(h_n_y+h_n_n))*math.log(h_n_y/(h_n_y+h_n_n),2))+((-h_n_n/(h_n_y+h_n_n))*math.log(h_n_n/(h_n_y+h_n_n),2))))
#print(info_hum)
infogain_hum = info_d-info_hum
#print(infogain_hum)

for row in range(1,sh.max_row+1):
    #col=6
    cellobj=sh.cell(row=row,column=5)
    cellobj2=sh.cell(row=row,column=6)
    #print(cellobj.value)
    if cellobj.value == "Strong" and cellobj2.value == "Yes":
        w_s_y += 1
    elif cellobj.value == "Strong" and cellobj2.value == "No":
        w_s_n +=1
    elif cellobj.value == "Weak" and cellobj2.value == "Yes":
        w_w_y += 1
    elif cellobj.value == "Weak" and cellobj2.value == "No":
        w_w_n += 1
#print(w_s_y,w_s_n,w_w_y,w_w_n)

info_wind = (((w_s_y+w_s_n)/(yc+nc))*(((-w_s_y/(w_s_y+w_s_n))*math.log(w_s_y/(w_s_y+w_s_n),2))+((-w_s_n/(w_s_y+w_s_n))*math.log(w_s_n/(w_s_y+w_s_n),2))))
+ (((w_w_y+w_w_n)/(yc+nc))*(((-w_w_y/(w_w_y+w_w_n))*math.log(w_w_y/(w_w_y+w_w_n),2))+((-w_w_n/(w_w_y+w_w_n))*math.log(w_w_n/(w_w_y+w_w_n),2))))

#print(info_wind)
infogain_wind = info_d-info_hum
#print(infogain_wind)

max1=max(infogain_hum,infogain_out,infogain_temp,infogain_temp,infogain_wind)
print(max1)
print(infogain_hum,infogain_out,infogain_temp,infogain_wind)